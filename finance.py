import pandas
import xlrd
import openpyxl
from datetime import datetime
from openpyxl import workbook,load_workbook
import time
from openpyxl.styles import numbers
from pathlib import Path
import sys
import os
from win32com.client import Dispatch

sys.argv
user_input = Path(input("Enter the path of your file: "))
f1 = pandas.read_excel(user_input,dtype=str,index_col=0)
f1.rename(columns={"Bank Account No":"Beneficiary Ac No","IFSC Code":"IFSC","Amount":"Amt"},inplace=True)
f1.insert(1,"Debit Ac No","182051000001")
f1.insert(5,"Pay Mod","N")
f1.insert(6,"Date","")
f1.insert(8,"Payable Loaction","")
f1.insert(9,"print Location","")
f1.insert(10,"Bene Mobile No","")
f1.insert(11,"Bene Email Id","")
f1.insert(12,"Bene add 1","")
f1.insert(13,"Bene add 2","")
f1.insert(14,"Bene add 3","")
f1.insert(15,"Bene add 4","")
f1.insert(16,"Add Details 1","")
f1.insert(17,"Add Details 2","")
f1.insert(18,"Add Details 3","")
f1.insert(19,"Add Details 4","")
f1.insert(20,"Add Details 5","")
f1.insert(21,"Remarks","")




DIR = os.path.dirname(user_input)
file = "output"+ os.path.basename(user_input) 

user_output = os.path.join(DIR, file)



first_column = f1.pop('Amt')
second_column = f1.pop('Beneficiary Name')
  
# insert column using insert(position,column_name,
# first_column) function
f1.insert(4, 'Amt', first_column)
f1.insert(4,'Beneficiary Name',second_column)

print('Final Excel sheet now generated at the same location:')

f1.to_excel(user_output, index = False)

wb = load_workbook(user_output)
sheet = wb.active
sheet.delete_cols(1)
sheet.delete_cols(2)
#sheet.delete_cols(7)
#sheet.insert_cols(idx =1)
#sheet.insert_cols(idx =5,amount=2)
#sheet.insert_cols(idx =8 , amount=14)



for rows in sheet.iter_rows(min_row=None, max_row=None, min_col=None, max_col=None):
    for cell in rows:
       cell.number_format =  numbers.FORMAT_TEXT
    wb.save(user_output)


#wb.save(user_output) 
xl = Dispatch('Excel.Application')
wb = xl.Workbooks.Add(user_output)
wb.SaveAs(user_output[:-1], FileFormat=56)
xl.Quit()


# "C:\Users\RITIKA ROY\Desktop\financeWork\Payroll_Bank_Statement (2).xlsx"


#C:\Users\RITIKA ROY\Desktop\financeWork\CMStestpay_PROACTIVEDATA .xls
#"C:\Users\RITIKA ROY\Desktop\financeWork\Payroll_Bank_Statement (2).xlsx"
#"C:\Users\RITIKA ROY\Desktop\financeWork\outputPayroll_Bank_Statement (2).xls"

